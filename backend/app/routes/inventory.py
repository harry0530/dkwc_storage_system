from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, schemas, crud
from html import escape
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import os
import re

router = APIRouter(prefix="/inventory", tags=["Inventory"])


FACTORY_LAYOUT_TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "templates",
    "factory_layout_template.xlsx",
)


def _normalize_location_code(value: str) -> str:
    raw = (value or "").strip().upper()
    if not raw:
        return ""
    match = re.match(r"^([A-Z])\s*-\s*([0-9]{1,3})$", raw)
    if match:
        zone = match.group(1)
        num = match.group(2)
        if len(num) == 1:
            num = f"0{num}"
        return f"{zone}-{num}"
    return re.sub(r"\s+", "", raw)


def _find_location_cell(workbook, location_code: str):
    target = _normalize_location_code(location_code)
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if _normalize_location_code(str(cell.value or "")) == target:
                    return ws, cell
    return None, None


def _fill_location_cell(ws, cell, fill):
    # Usually location codes are single cells, but this keeps merged cells safe too.
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            for row in ws[merged_range.coord]:
                for merged_cell in row:
                    merged_cell.fill = fill
            return
    cell.fill = fill


def _excel_col_width_to_px(width):
    return max(24, int(round((width or 8.43) * 7 + 5)))


def _excel_row_height_to_px(height):
    return max(18, int(round((height or 15) * 96 / 72)))


def _theme_color(theme_index):
    theme_map = {
        0: "FFFFFF",
        1: "000000",
        2: "EEECE1",
        3: "1F497D",
        4: "4F81BD",
        5: "FCE4D6",
        6: "9BBB59",
        7: "8064A2",
        8: "4BACC6",
        9: "F79646",
    }
    return theme_map.get(theme_index)


def _style_color_to_hex(color):
    if not color:
        return None
    if color.type == "rgb" and isinstance(color.rgb, str):
        rgb = color.rgb[-6:].upper()
        if color.rgb.upper() == "00000000":
            return None
        return rgb
    if color.type == "theme":
        return _theme_color(color.theme)
    return None


def _cell_fill_hex(cell):
    if not cell.fill or cell.fill.fill_type != "solid":
        return None
    return _style_color_to_hex(cell.fill.fgColor)


def _font_color_hex(cell):
    color = getattr(cell.font, "color", None)
    return _style_color_to_hex(color) or "0F172A"


def _border_width(style):
    if not style:
        return 0
    return 2 if style in {"medium", "thick", "double"} else 1


def _has_visible_cell(cell):
    if cell.value not in (None, ""):
        return True
    if _cell_fill_hex(cell):
        return True
    border = cell.border
    return any(
        getattr(border, side).style
        for side in ("left", "right", "top", "bottom")
    )


def _make_sheet_metrics(ws):
    col_widths = [
        _excel_col_width_to_px(ws.column_dimensions[get_column_letter(col)].width)
        for col in range(1, ws.max_column + 1)
    ]
    row_heights = [
        _excel_row_height_to_px(ws.row_dimensions[row].height)
        for row in range(1, ws.max_row + 1)
    ]

    col_lefts = [0]
    for width in col_widths:
        col_lefts.append(col_lefts[-1] + width)

    row_tops = [0]
    for height in row_heights:
        row_tops.append(row_tops[-1] + height)

    return col_widths, row_heights, col_lefts, row_tops


def _cell_rect(bounds, col_lefts, row_tops):
    min_col, min_row, max_col, max_row = bounds
    x = col_lefts[min_col - 1]
    y = row_tops[min_row - 1]
    w = col_lefts[max_col] - x
    h = row_tops[max_row] - y
    return x, y, w, h


def _build_location_map_svg(ws, target_cell, danger=False):
    col_widths, row_heights, col_lefts, row_tops = _make_sheet_metrics(ws)
    width = sum(col_widths)
    height = sum(row_heights)

    merged_starts = {}
    merged_children = set()
    target_bounds = (
        target_cell.column,
        target_cell.row,
        target_cell.column,
        target_cell.row,
    )

    for merged_range in ws.merged_cells.ranges:
        bounds = (
            merged_range.min_col,
            merged_range.min_row,
            merged_range.max_col,
            merged_range.max_row,
        )
        merged_starts[(merged_range.min_row, merged_range.min_col)] = bounds
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (merged_range.min_row, merged_range.min_col):
                    merged_children.add((row, col))
        if target_cell.coordinate in merged_range:
            target_bounds = bounds

    fill_parts = [
        f'<rect x="0" y="0" width="{width}" height="{height}" fill="#ffffff"/>'
    ]
    border_parts = []
    text_parts = []

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if (row, col) in merged_children:
                continue

            bounds = merged_starts.get((row, col), (col, row, col, row))
            cell = ws.cell(row=row, column=col)
            if not _has_visible_cell(cell) and bounds != target_bounds:
                continue

            x, y, w, h = _cell_rect(bounds, col_lefts, row_tops)
            fill = "F4CCCC" if danger and bounds == target_bounds else None
            if bounds == target_bounds and not fill:
                fill = "FFF2CC"
            fill = fill or _cell_fill_hex(cell) or "FFFFFF"
            fill_parts.append(
                f'<rect x="{x}" y="{y}" width="{w}" height="{h}" fill="#{fill}"/>'
            )

            border = cell.border
            for side, x1, y1, x2, y2 in (
                ("top", x, y, x + w, y),
                ("right", x + w, y, x + w, y + h),
                ("bottom", x, y + h, x + w, y + h),
                ("left", x, y, x, y + h),
            ):
                border_side = getattr(border, side)
                line_width = _border_width(border_side.style)
                if line_width:
                    border_color = _style_color_to_hex(border_side.color) or "222222"
                    border_parts.append(
                        f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" '
                        f'stroke="#{border_color}" stroke-width="{line_width}" '
                        'shape-rendering="crispEdges"/>'
                    )

            if cell.value not in (None, ""):
                text = escape(str(cell.value))
                font_size = float(cell.font.sz or 11) * 1.25
                font_weight = "700" if cell.font.bold else "500"
                text_color = _font_color_hex(cell)
                text_parts.append(
                    f'<text x="{x + w / 2}" y="{y + h / 2}" '
                    f'text-anchor="middle" dominant-baseline="central" '
                    f'font-size="{font_size:.1f}" font-weight="{font_weight}" '
                    f'fill="#{text_color}">{text}</text>'
                )

    # Draw a final highlight overlay on top so it never gets lost.
    hx, hy, hw, hh = _cell_rect(target_bounds, col_lefts, row_tops)
    highlight_fill = "F4CCCC" if danger else "FFF2CC"
    highlight_stroke = "B91C1C" if danger else "CA8A04"
    highlight_parts = [
        f'<rect x="{hx}" y="{hy}" width="{hw}" height="{hh}" fill="#{highlight_fill}" fill-opacity="0.55"/>',
        f'<rect x="{hx}" y="{hy}" width="{hw}" height="{hh}" fill="none" stroke="#{highlight_stroke}" stroke-width="3"/>',
    ]

    return (
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {width} {height}" '
        f'width="{width}" height="{height}">'
        '<style>text{font-family:"Malgun Gothic","Apple SD Gothic Neo",Arial,sans-serif;}</style>'
        + "".join(fill_parts)
        + "".join(border_parts)
        + "".join(text_parts)
        + "".join(highlight_parts)
        + "</svg>"
    )


# ✅ 재고 조회 (JOIN 버전)
@router.get("/")
def get_inventory(db: Session = Depends(get_db)):
    result = []

    inventory_list = db.query(models.Product).filter(
        models.Product.type == "PART"
    ).all()

    for inv in inventory_list:
        result.append({
            "code": inv.new_code,
            "product_code": inv.new_code,
            "old_code": inv.old_code,
            "new_code": inv.new_code,
            "drawing_number": inv.drawing_number or "",
            "name": inv.name or "",
            "material": inv.material or "",
            "spec": inv.spec or "",
            "heat_treatment": inv.heat_treatment or "",
            "welding": inv.welding or "",
            "plating": inv.plating or "",
            "type": inv.type,
            "location": inv.location or "",
            "quantity": inv.quantity,
            "min_stock": inv.min_stock,
            "supplier_company_id": inv.supplier_company_id
        })

    return result


@router.get("/location-map-image/{location_code}")
def view_location_map_image(location_code: str, danger: bool = False):
    normalized_code = _normalize_location_code(location_code)
    if not normalized_code:
        raise HTTPException(status_code=400, detail="보관위치를 입력하세요.")
    if not os.path.exists(FACTORY_LAYOUT_TEMPLATE):
        raise HTTPException(status_code=500, detail="공장 배치도 템플릿을 찾을 수 없습니다.")

    workbook = load_workbook(FACTORY_LAYOUT_TEMPLATE)
    ws, cell = _find_location_cell(workbook, normalized_code)
    if not cell:
        raise HTTPException(status_code=404, detail=f"배치도에서 {normalized_code} 위치를 찾을 수 없습니다.")

    svg = _build_location_map_svg(ws, cell, danger=danger)
    return Response(content=svg, media_type="image/svg+xml")


@router.get("/location-map/{location_code}")
def download_location_map(location_code: str, danger: bool = False):
    normalized_code = _normalize_location_code(location_code)
    if not normalized_code:
        raise HTTPException(status_code=400, detail="보관위치를 입력하세요.")
    if not os.path.exists(FACTORY_LAYOUT_TEMPLATE):
        raise HTTPException(status_code=500, detail="공장 배치도 템플릿을 찾을 수 없습니다.")

    workbook = load_workbook(FACTORY_LAYOUT_TEMPLATE)
    ws, cell = _find_location_cell(workbook, normalized_code)
    if not cell:
        raise HTTPException(status_code=404, detail=f"배치도에서 {normalized_code} 위치를 찾을 수 없습니다.")

    fill_color = "F4CCCC" if danger else "FFF2CC"
    _fill_location_cell(ws, cell, PatternFill(fill_type="solid", fgColor=fill_color))
    workbook.active = workbook.worksheets.index(ws)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"factory_layout_{normalized_code}.xlsx"
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ✅ 재고 입력 (복구됨 ⭐)
@router.post("/")
def create_inventory(inv: schemas.InventoryCreate, db: Session = Depends(get_db)):
    return crud.create_inventory(db, inv)


# ✅ 재고 수정 (수량 직접 수정)
@router.put("/{product_code}")
def update_inventory(product_code: str, data: dict, db: Session = Depends(get_db)):
    inv = db.query(models.Product).filter(
        models.Product.new_code == product_code
    ).first()

    if not inv:
        raise HTTPException(status_code=404, detail="재고 없음")

    if "quantity" in data:
        new_qty = data["quantity"]
        prev_qty = inv.quantity
        inv.quantity = new_qty

        diff = new_qty - prev_qty
        if diff != 0:
            reason = data.get("reason") or "수량 변경"
            tx_type = "IN" if diff > 0 else "OUT"
            db.add(models.Transaction(
                product_code=inv.new_code,
                quantity=abs(diff),
                type=tx_type,
                reason=reason
            ))

    db.commit()

    return {"message": "수정 완료"}


# ✅ 재고 삭제 (목록에서 제거)
@router.delete("/{product_code}")
def delete_inventory(product_code: str, db: Session = Depends(get_db)):
    inv = db.query(models.Product).filter(
        models.Product.new_code == product_code
    ).first()

    if not inv:
        raise HTTPException(status_code=404, detail="재고 없음")

    db.delete(inv)
    db.commit()

    return {"message": "삭제 완료"}

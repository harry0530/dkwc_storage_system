from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import or_
from io import BytesIO
import openpyxl
from app.database import get_db
from app import models, schemas

router = APIRouter(prefix="/products", tags=["Products"])


PART_IMPORT_ALIASES = {
    "old_code": ["기존품번", "구품번"],
    "new_code": ["신품번"],
    "drawing_number": ["도번", "drawing_number"],
    "name": ["품명"],
    "material": ["재질"],
    "spec": ["규격"],
    "quantity": ["현재재고", "재고수량", "재고"],
    "min_stock": ["최소재고"],
    "location": ["보관위치"],
    "supplier_name": ["납품처", "발주처"],
}


def _normalize_excel_value(value):
    return str(value or "").strip()


def _normalize_excel_code(value):
    raw = _normalize_excel_value(value)
    if raw.endswith(".0"):
        trimmed = raw[:-2]
        if trimmed.isdigit():
            return trimmed
    return raw


def _to_int(value):
    text = _normalize_excel_value(value)
    if not text:
        return 0
    try:
        return int(float(text.replace(",", "")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"숫자 형식 오류: {text}") from exc


def _build_part_header_map(values):
    header_values = [_normalize_excel_value(v) for v in values]
    raw_mapping = {name: idx for idx, name in enumerate(header_values) if name}
    canonical_mapping = {}
    for canonical_name, aliases in PART_IMPORT_ALIASES.items():
        for alias in aliases:
            if alias in raw_mapping:
                canonical_mapping[canonical_name] = raw_mapping[alias]
                break
    return header_values, canonical_mapping


def _find_parts_sheet(workbook):
    selected_ws = workbook.active
    selected_header = []
    selected_col = {}
    header_row = 1

    for sheet_name in workbook.sheetnames:
        candidate = workbook[sheet_name]
        max_rows = min(candidate.max_row, 10)
        for row_idx in range(1, max_rows + 1):
            header, col = _build_part_header_map(
                [cell.value for cell in candidate[row_idx]]
            )
            if "new_code" in col:
                return candidate, header, col, row_idx

    return selected_ws, selected_header, selected_col, header_row


# 🔥 핵심 수정 부분
@router.post("/")
def create_product(product: schemas.ProductCreate, db: Session = Depends(get_db)):
    code = product.code.strip()
    name = product.name.strip() if product.name else ""

    existing = db.query(models.Product).filter(
        models.Product.new_code == code
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="이미 존재하는 품번")

    db_product = models.Product(
        old_code=(product.old_code or "").strip(),
        new_code=code,
        drawing_number=(product.drawing_number or "").strip(),
        name=name,
        type=product.type,
        material=(product.material or "").strip(),
        spec=(product.spec or "").strip(),
        quantity=product.quantity or 0,
        location=(product.location or "").strip(),
        min_stock=product.min_stock,
        supplier_company_id=product.supplier_company_id
    )

    db.add(db_product)
    db.commit()
    db.refresh(db_product)

    return db_product


@router.post("/import-parts")
def import_parts(
    file: UploadFile = File(...),
    duplicate_action: str = Form("prompt"),
    db: Session = Depends(get_db)
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="파일이 없습니다")

    duplicate_action = (duplicate_action or "prompt").strip().lower()
    if duplicate_action not in {"prompt", "overwrite", "skip"}:
        raise HTTPException(status_code=400, detail="중복 처리 방식이 올바르지 않습니다")

    content = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws, header, col, header_row = _find_parts_sheet(wb)

    required = [
        "old_code",
        "new_code",
        "name",
        "spec",
        "material",
        "quantity",
        "min_stock",
        "location",
        "supplier_name",
    ]
    missing = [field for field in required if field not in col]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"엑셀 형식 오류: 필요한 컬럼이 없습니다 ({', '.join(missing)})"
        )

    parsed_rows = []
    duplicate_codes = []
    rows_total = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not any(value not in (None, "") for value in row):
            continue

        rows_total += 1
        new_code = _normalize_excel_code(row[col["new_code"]])
        if not new_code:
            parsed_rows.append({"is_blank": True})
            continue

        parsed = {
            "old_code": _normalize_excel_code(row[col["old_code"]]),
            "new_code": new_code,
            "drawing_number": _normalize_excel_value(row[col["drawing_number"]]) if "drawing_number" in col else "",
            "name": _normalize_excel_value(row[col["name"]]),
            "spec": _normalize_excel_value(row[col["spec"]]),
            "material": _normalize_excel_value(row[col["material"]]),
            "quantity": _to_int(row[col["quantity"]]),
            "min_stock": _to_int(row[col["min_stock"]]),
            "location": _normalize_excel_value(row[col["location"]]),
            "supplier_name": _normalize_excel_value(row[col["supplier_name"]]),
        }
        parsed_rows.append(parsed)

        product = db.query(models.Product).filter(
            models.Product.new_code == parsed["new_code"]
        ).first()
        if product:
            duplicate_codes.append(parsed["new_code"])

    if duplicate_codes and duplicate_action == "prompt":
        unique_duplicates = list(dict.fromkeys(duplicate_codes))
        raise HTTPException(
            status_code=409,
            detail={
                "message": "이미 등록된 단품이 있습니다.",
                "duplicate_count": len(unique_duplicates),
                "duplicate_codes": unique_duplicates[:20],
                "rows_total": rows_total,
            },
        )

    created = 0
    updated = 0
    skipped = 0

    for parsed in parsed_rows:
        if parsed.get("is_blank"):
            skipped += 1
            continue

        supplier_id = None
        if parsed["supplier_name"]:
            company = db.query(models.Company).filter(
                models.Company.name == parsed["supplier_name"]
            ).first()
            if not company:
                company = models.Company(
                    name=parsed["supplier_name"], phone="", fax="", address=""
                )
                db.add(company)
                db.commit()
                db.refresh(company)
            supplier_id = company.id

        product = db.query(models.Product).filter(
            models.Product.new_code == parsed["new_code"]
        ).first()

        if product:
            if duplicate_action == "skip":
                skipped += 1
                continue

            product.old_code = parsed["old_code"]
            product.drawing_number = parsed["drawing_number"]
            product.name = parsed["name"]
            product.type = "PART"
            product.material = parsed["material"]
            product.spec = parsed["spec"]
            product.quantity = parsed["quantity"]
            product.min_stock = parsed["min_stock"]
            product.location = parsed["location"]
            product.supplier_company_id = supplier_id
            updated += 1
        else:
            db.add(models.Product(
                old_code=parsed["old_code"],
                new_code=parsed["new_code"],
                drawing_number=parsed["drawing_number"],
                name=parsed["name"],
                type="PART",
                material=parsed["material"],
                spec=parsed["spec"],
                quantity=parsed["quantity"],
                min_stock=parsed["min_stock"],
                location=parsed["location"],
                supplier_company_id=supplier_id
            ))
            created += 1

    db.commit()

    sample_rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=True):
        sample_rows.append([_normalize_excel_value(value) for value in row])

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "rows_total": rows_total,
        "duplicate_action": duplicate_action,
        "sheet": ws.title,
        "header": header,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "header_row": header_row,
        "sample_rows": sample_rows
    }


@router.post("/import-finished")
def import_finished(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename:
        raise HTTPException(status_code=400, detail="파일이 없습니다")

    content = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(header)}

    required = ["기존품번", "신품번", "품명", "규격", "재질", "BOM", "발주처"]
    for r in required:
        if r not in col:
            raise HTTPException(status_code=400, detail=f"엑셀 형식 오류: {r} 컬럼 없음")

    created = 0
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        old_code = str(row[col["기존품번"]] or "").strip()
        new_code = str(row[col["신품번"]] or "").strip()
        drawing_number = str(row[col["도번"]] or "").strip() if "도번" in col else ""
        name = str(row[col["품명"]] or "").strip()
        spec = str(row[col["규격"]] or "").strip()
        material = str(row[col["재질"]] or "").strip()
        bom_text = (row[col["BOM"]] or "").strip()
        supplier_name = str(row[col["발주처"]] or "").strip()

        if not new_code:
            continue

        supplier_id = None
        if supplier_name:
            company = db.query(models.Company).filter(
                models.Company.name == supplier_name
            ).first()
            if not company:
                company = models.Company(name=supplier_name, phone="", fax="", address="")
                db.add(company)
                db.commit()
                db.refresh(company)
            supplier_id = company.id

        product = db.query(models.Product).filter(
            models.Product.new_code == new_code
        ).first()

        if product:
            product.old_code = old_code
            product.drawing_number = drawing_number
            product.name = name
            product.type = "FINISHED"
            product.material = material
            product.spec = spec
            product.supplier_company_id = supplier_id
            updated += 1
        else:
            db.add(models.Product(
                old_code=old_code,
                new_code=new_code,
                drawing_number=drawing_number,
                name=name,
                type="FINISHED",
                material=material,
                spec=spec,
                quantity=0,
                min_stock=0,
                location="",
                supplier_company_id=supplier_id
            ))
            created += 1

        if bom_text:
            for chunk in bom_text.split(","):
                part = chunk.strip()
                if not part or ":" not in part:
                    continue
                child_code, qty_text = part.split(":", 1)
                child_code = child_code.strip()
                try:
                    qty = int(qty_text.strip())
                except Exception:
                    continue

                existing = db.query(models.BOM).filter(
                    models.BOM.parent_code == new_code,
                    models.BOM.child_code == child_code
                ).first()
                if not existing:
                    db.add(models.BOM(
                        parent_code=new_code,
                        child_code=child_code,
                        quantity=qty
                    ))

    db.commit()

    return {"created": created, "updated": updated}


@router.get("/")
def get_products(db: Session = Depends(get_db)):
    return db.query(models.Product).all()


@router.delete("/{product_code}")
def delete_product(product_code: str, db: Session = Depends(get_db)):
    product = db.query(models.Product).filter(
        models.Product.new_code == product_code
    ).first()

    if not product:
        raise HTTPException(status_code=404, detail="제품 없음")

    # 제품 삭제 시 참조되는 기본 데이터 정리
    db.query(models.BOM).filter(
        or_(
            models.BOM.parent_code == product_code,
            models.BOM.child_code == product_code
        )
    ).delete(synchronize_session=False)

    db.delete(product)
    db.commit()

    return {"message": "삭제 완료"}


# ⭐ 제품 수정
@router.put("/{product_code}")
def update_product(product_code: str, data: dict, db: Session = Depends(get_db)):
    product = db.query(models.Product).filter(
        models.Product.new_code == product_code
    ).first()

    if not product:
        raise HTTPException(status_code=404, detail="제품 없음")

    product.old_code = data.get("old_code", product.old_code)
    product.drawing_number = data.get("drawing_number", product.drawing_number)
    product.name = data.get("name", product.name)
    product.type = data.get("type", product.type)
    product.material = data.get("material", product.material)
    product.spec = data.get("spec", product.spec)
    product.location = data.get("location", product.location)
    product.min_stock = data.get("min_stock", product.min_stock)
    product.quantity = data.get("quantity", product.quantity)
    product.supplier_company_id = data.get(
        "supplier_company_id", product.supplier_company_id
    )

    db.commit()
    return product


@router.put("/{product_code}/change-code")
def change_product_code(product_code: str, data: dict, db: Session = Depends(get_db)):
    old_code = (product_code or "").strip()
    new_code = (data.get("new_code") or "").strip()

    if not old_code:
        raise HTTPException(status_code=400, detail="기존 신품번이 필요합니다")
    if not new_code:
        raise HTTPException(status_code=400, detail="새 신품번을 입력하세요")
    if old_code == new_code:
        raise HTTPException(status_code=400, detail="같은 신품번으로는 변경할 수 없습니다")

    product = db.query(models.Product).filter(
        models.Product.new_code == old_code
    ).first()
    if not product:
        raise HTTPException(status_code=404, detail="제품 없음")

    duplicate = db.query(models.Product).filter(
        models.Product.new_code == new_code
    ).first()
    if duplicate:
        raise HTTPException(status_code=400, detail="이미 존재하는 신품번입니다")

    # Product references
    product.new_code = new_code

    # BOM references
    db.query(models.BOM).filter(
        models.BOM.parent_code == old_code
    ).update({models.BOM.parent_code: new_code}, synchronize_session=False)
    db.query(models.BOM).filter(
        models.BOM.child_code == old_code
    ).update({models.BOM.child_code: new_code}, synchronize_session=False)

    # Order / purchase references
    db.query(models.Order).filter(
        models.Order.product_code == old_code
    ).update({models.Order.product_code: new_code}, synchronize_session=False)
    db.query(models.PurchaseOrder).filter(
        models.PurchaseOrder.product_code == old_code
    ).update({models.PurchaseOrder.product_code: new_code}, synchronize_session=False)

    # Transaction history references
    db.query(models.Transaction).filter(
        models.Transaction.product_code == old_code
    ).update({models.Transaction.product_code: new_code}, synchronize_session=False)

    # Optional alias references when that model exists in this deployment
    alias_model = getattr(models, "ProductAlias", None)
    if alias_model is not None:
        db.query(alias_model).filter(
            alias_model.product_code == old_code
        ).update({alias_model.product_code: new_code}, synchronize_session=False)

    db.commit()

    return {
        "message": "신품번 변경 완료",
        "old_code": old_code,
        "new_code": new_code,
    }

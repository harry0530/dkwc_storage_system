from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app import crud, models, schemas
from app.database import get_db

router = APIRouter(prefix="/bom", tags=["BOM"])

BLOCK_SECTION_TITLE = "대상 완제품"
PART_SECTION_TITLE = "대상 단품"

PARENT_HEADER_ALIASES = {
    "old_code": ["완제품 구품번", "구품번"],
    "new_code": ["완제품 신품번", "신품번"],
    "name": ["품명"],
    "spec": ["규격", "규격/비고"],
}

PART_HEADER_ALIASES = {
    "old_code": ["부품 구품번", "구품번"],
    "new_code": ["부품 신품번", "신품번"],
    "name": ["품명"],
    "spec": ["규격"],
    "material": ["재질"],
    "quantity": ["소요량"],
    "supplier": ["업체", "납품처"],
}


def _normalize_excel_value(value):
    return str(value or "").strip()


def _normalize_excel_code(value):
    raw = _normalize_excel_value(value)
    if raw.endswith(".0"):
        trimmed = raw[:-2]
        if trimmed.isdigit():
            return trimmed
    return raw


def _to_int(value):
    text = _normalize_excel_value(value)
    if not text:
        return 0
    try:
        return int(float(text.replace(",", "")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"숫자 형식 오류: {text}") from exc


def _build_header_map(values, aliases):
    header_values = [_normalize_excel_value(v) for v in values]
    raw_mapping = {name: idx for idx, name in enumerate(header_values) if name}
    canonical_mapping = {}
    for canonical_name, alias_list in aliases.items():
        for alias in alias_list:
            if alias in raw_mapping:
                canonical_mapping[canonical_name] = raw_mapping[alias]
                break
    return canonical_mapping


def _find_product(db: Session, *, new_code="", old_code="", product_type=None):
    filters = []
    if new_code:
        filters.append(models.Product.new_code == new_code)
    if old_code:
        filters.append(models.Product.old_code == old_code)
    if not filters:
        return None

    query = db.query(models.Product)
    if product_type:
        query = query.filter(models.Product.type == product_type)
    return query.filter(or_(*filters)).first()


def _generate_auto_finished_code(db: Session, reserved_codes: set[str]):
    max_numeric = -1

    existing_codes = db.query(models.Product.new_code).filter(
        models.Product.type == "FINISHED"
    ).all()
    for (existing_code,) in existing_codes:
        code = (existing_code or "").strip()
        if code.isdigit():
            max_numeric = max(max_numeric, int(code))

    for code in reserved_codes:
        if code.isdigit():
            max_numeric = max(max_numeric, int(code))

    next_number = max_numeric + 1
    next_code = str(next_number).zfill(4)
    while next_code in reserved_codes:
        next_number += 1
        next_code = str(next_number).zfill(4)

    reserved_codes.add(next_code)
    return next_code


def _parse_block_sheet(ws):
    rows = [
        [cell for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]

    parsed_blocks = []
    row_index = 0

    while row_index < len(rows):
        first_cell = _normalize_excel_value(rows[row_index][0] if rows[row_index] else "")
        if first_cell != BLOCK_SECTION_TITLE:
            row_index += 1
            continue

        if row_index + 2 >= len(rows):
            raise HTTPException(status_code=400, detail="완제품 블록 형식이 올바르지 않습니다.")

        parent_header_row = rows[row_index + 1]
        parent_value_row = rows[row_index + 2]
        parent_col = _build_header_map(parent_header_row, PARENT_HEADER_ALIASES)

        if "old_code" not in parent_col and "new_code" not in parent_col:
            raise HTTPException(status_code=400, detail="완제품 코드 헤더를 찾을 수 없습니다.")

        part_section_index = None
        for candidate_index in range(row_index + 3, len(rows)):
            candidate_label = _normalize_excel_value(
                rows[candidate_index][0] if rows[candidate_index] else ""
            )
            if candidate_label == PART_SECTION_TITLE:
                part_section_index = candidate_index
                break
            if candidate_label == BLOCK_SECTION_TITLE:
                break

        if part_section_index is None or part_section_index + 1 >= len(rows):
            raise HTTPException(status_code=400, detail="대상 단품 영역을 찾을 수 없습니다.")

        part_header_row = rows[part_section_index + 1]
        part_col = _build_header_map(part_header_row, PART_HEADER_ALIASES)
        if "quantity" not in part_col:
            raise HTTPException(status_code=400, detail="소요량 헤더를 찾을 수 없습니다.")

        next_block_index = len(rows)
        for candidate_index in range(part_section_index + 2, len(rows)):
            candidate_label = _normalize_excel_value(
                rows[candidate_index][0] if rows[candidate_index] else ""
            )
            if candidate_label == BLOCK_SECTION_TITLE:
                next_block_index = candidate_index
                break

        parent = {
            "old_code": _normalize_excel_code(
                parent_value_row[parent_col["old_code"]] if "old_code" in parent_col else ""
            ),
            "new_code": _normalize_excel_code(
                parent_value_row[parent_col["new_code"]] if "new_code" in parent_col else ""
            ),
            "name": _normalize_excel_value(
                parent_value_row[parent_col["name"]] if "name" in parent_col else ""
            ),
            "spec": _normalize_excel_value(
                parent_value_row[parent_col["spec"]] if "spec" in parent_col else ""
            ),
        }

        if not parent["old_code"] and not parent["new_code"]:
            raise HTTPException(status_code=400, detail="완제품 코드가 비어 있습니다.")

        children = []
        for data_index in range(part_section_index + 2, next_block_index):
            row = rows[data_index]
            if not row or not any(value not in (None, "") for value in row):
                continue

            child_old_code = _normalize_excel_code(
                row[part_col["old_code"]] if "old_code" in part_col else ""
            )
            child_new_code = _normalize_excel_code(
                row[part_col["new_code"]] if "new_code" in part_col else ""
            )
            child_name = _normalize_excel_value(
                row[part_col["name"]] if "name" in part_col else ""
            )
            child_spec = _normalize_excel_value(
                row[part_col["spec"]] if "spec" in part_col else ""
            )
            child_material = _normalize_excel_value(
                row[part_col["material"]] if "material" in part_col else ""
            )
            child_supplier = _normalize_excel_value(
                row[part_col["supplier"]] if "supplier" in part_col else ""
            )
            quantity = _to_int(row[part_col["quantity"]])

            has_child_data = any(
                [child_old_code, child_new_code, child_name, child_spec, child_material, child_supplier]
            )
            if not has_child_data and quantity == 0:
                continue

            if not child_old_code and not child_new_code:
                raise HTTPException(
                    status_code=400,
                    detail=f"{parent['old_code'] or parent['new_code']} 블록에 부품 코드가 비어 있습니다.",
                )
            if quantity <= 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"{child_old_code or child_new_code} 소요량이 올바르지 않습니다.",
                )

            children.append(
                {
                    "old_code": child_old_code,
                    "new_code": child_new_code,
                    "quantity": quantity,
                }
            )

        if not children:
            raise HTTPException(
                status_code=400,
                detail=f"{parent['old_code'] or parent['new_code']} 블록에 등록할 단품이 없습니다.",
            )

        parsed_blocks.append({"parent": parent, "children": children})
        row_index = next_block_index

    if not parsed_blocks:
        raise HTTPException(status_code=400, detail="블록형 BOM 데이터를 찾지 못했습니다.")

    return parsed_blocks


@router.post("/")
def create_bom(bom: schemas.BOMCreate, db: Session = Depends(get_db)):
    return crud.create_bom(db, bom)


@router.post("/import-blocks")
def import_bom_blocks(
    file: UploadFile = File(...),
    duplicate_action: str = Form("prompt"),
    db: Session = Depends(get_db),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="파일이 없습니다.")

    duplicate_action = (duplicate_action or "prompt").strip().lower()
    if duplicate_action not in {"prompt", "overwrite", "skip"}:
        raise HTTPException(status_code=400, detail="중복 처리 방식이 올바르지 않습니다.")

    content = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb["블록형 BOM 양식"] if "블록형 BOM 양식" in wb.sheetnames else wb.active
    parsed_blocks = _parse_block_sheet(ws)

    reserved_codes = {
        (code or "").strip()
        for (code,) in db.query(models.Product.new_code).all()
        if (code or "").strip()
    }

    duplicate_parents = []
    prepared_blocks = []
    for block in parsed_blocks:
        parent_data = block["parent"]
        parent_product = _find_product(
            db,
            new_code=parent_data["new_code"],
            old_code=parent_data["old_code"],
            product_type="FINISHED",
        )

        target_code = parent_product.new_code if parent_product else parent_data["new_code"]
        has_existing_bom = bool(
            target_code
            and db.query(models.BOM).filter(models.BOM.parent_code == target_code).first()
        )
        if has_existing_bom:
            duplicate_parents.append(target_code)

        prepared_blocks.append(
            {
                "parent": parent_data,
                "parent_product": parent_product,
                "has_existing_bom": has_existing_bom,
                "children": block["children"],
            }
        )

    if duplicate_parents and duplicate_action == "prompt":
        unique_duplicates = list(dict.fromkeys(duplicate_parents))
        raise HTTPException(
            status_code=409,
            detail={
                "message": "이미 BOM이 등록된 완제품이 있습니다.",
                "duplicate_count": len(unique_duplicates),
                "duplicate_codes": unique_duplicates[:20],
                "blocks_total": len(parsed_blocks),
            },
        )

    blocks_total = len(parsed_blocks)
    created_finished = 0
    updated_finished = 0
    overwritten_blocks = 0
    skipped_blocks = 0
    bom_rows_total = 0
    created_bom_rows = 0

    for block in prepared_blocks:
        parent_data = block["parent"]
        parent_product = block["parent_product"]

        if block["has_existing_bom"] and duplicate_action == "skip":
            skipped_blocks += 1
            continue

        if parent_product:
            parent_code = parent_product.new_code
            parent_product.old_code = parent_data["old_code"]
            parent_product.name = parent_data["name"] or parent_product.name
            parent_product.spec = parent_data["spec"] or parent_product.spec
            parent_product.type = "FINISHED"
            parent_product.material = ""
            parent_product.quantity = 0
            parent_product.min_stock = 0
            parent_product.location = ""
            parent_product.supplier_company_id = None
            updated_finished += 1
        else:
            parent_code = parent_data["new_code"] or _generate_auto_finished_code(db, reserved_codes)
            db.add(
                models.Product(
                    old_code=parent_data["old_code"],
                    new_code=parent_code,
                    drawing_number="",
                    name=parent_data["name"],
                    type="FINISHED",
                    material="",
                    spec=parent_data["spec"],
                    quantity=0,
                    min_stock=0,
                    location="",
                    supplier_company_id=None,
                )
            )
            created_finished += 1

        if block["has_existing_bom"] and duplicate_action == "overwrite":
            db.query(models.BOM).filter(
                models.BOM.parent_code == parent_code
            ).delete(synchronize_session=False)
            overwritten_blocks += 1

        seen_child_codes = set()
        for child in block["children"]:
            child_product = _find_product(
                db,
                new_code=child["new_code"],
                old_code=child["old_code"],
                product_type="PART",
            )
            if not child_product:
                raise HTTPException(
                    status_code=400,
                    detail=f"부품을 찾을 수 없습니다: {child['new_code'] or child['old_code']}",
                )

            child_code = child_product.new_code
            if child_code in seen_child_codes:
                existing = db.query(models.BOM).filter(
                    models.BOM.parent_code == parent_code,
                    models.BOM.child_code == child_code,
                ).first()
                if existing:
                    existing.quantity = child["quantity"]
                continue

            seen_child_codes.add(child_code)
            bom_rows_total += 1
            existing = db.query(models.BOM).filter(
                models.BOM.parent_code == parent_code,
                models.BOM.child_code == child_code,
            ).first()
            if existing:
                existing.quantity = child["quantity"]
            else:
                db.add(
                    models.BOM(
                        parent_code=parent_code,
                        child_code=child_code,
                        quantity=child["quantity"],
                    )
                )
                created_bom_rows += 1

    db.commit()

    return {
        "blocks_total": blocks_total,
        "created_finished": created_finished,
        "updated_finished": updated_finished,
        "overwritten_blocks": overwritten_blocks,
        "skipped_blocks": skipped_blocks,
        "bom_rows_total": bom_rows_total,
        "created_bom_rows": created_bom_rows,
        "duplicate_action": duplicate_action,
    }


@router.get("/{product_code}")
def get_bom(product_code: str, db: Session = Depends(get_db)):
    return crud.get_bom_by_product(db, product_code)


@router.get("/")
def get_all_bom(db: Session = Depends(get_db)):
    return crud.get_all_bom(db)


@router.delete("/{bom_id}")
def delete_bom(bom_id: int, db: Session = Depends(get_db)):
    return crud.delete_bom(db, bom_id)


@router.put("/{bom_id}")
def update_bom(bom_id: int, data: schemas.BOMUpdate, db: Session = Depends(get_db)):
    return crud.update_bom(db, bom_id, data)

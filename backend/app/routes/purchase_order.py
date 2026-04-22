from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, schemas
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from typing import Optional

import openpyxl
import zipfile
import xml.etree.ElementTree as ET

router = APIRouter(prefix="/purchase-orders", tags=["PurchaseOrders"])

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates" / "purchase_order_template.xlsx"

REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _pick_free_rid(existing: set[str], start: int = 2) -> str:
    i = start
    while True:
        rid = f"rId{i}"
        if rid not in existing:
            return rid
        i += 1


def _merge_content_types(
    *,
    out_xml: bytes,
    template_xml: bytes,
    extra_overrides: dict[str, str] | None = None,
) -> bytes:
    # openpyxl drops drawings/media parts, which also removes Content_Types entries.
    # We merge template entries back so Office can load the embedded images/VML.
    ET.register_namespace("", "http://schemas.openxmlformats.org/package/2006/content-types")
    out_root = ET.fromstring(out_xml)
    tmpl_root = ET.fromstring(template_xml)

    def _key(elem):
        tag = elem.tag.rsplit("}", 1)[-1]
        if tag == "Default":
            return ("Default", elem.get("Extension"))
        if tag == "Override":
            return ("Override", elem.get("PartName"))
        return (tag, None)

    existing = {_key(e) for e in out_root}
    for e in tmpl_root:
        k = _key(e)
        if k in existing:
            continue
        out_root.append(e)
        existing.add(k)

    if extra_overrides:
        for part_name, content_type in extra_overrides.items():
            k = ("Override", part_name)
            if k in existing:
                continue
            ET.SubElement(
                out_root,
                "{http://schemas.openxmlformats.org/package/2006/content-types}Override",
                PartName=part_name,
                ContentType=content_type,
            )
            existing.add(k)

    return ET.tostring(out_root, encoding="utf-8", xml_declaration=True)


def _ensure_sheet_has_drawings(*, sheet_xml: bytes, drawing_rid: str, vml_rid: str, vmlhf_rid: str) -> bytes:
    text = sheet_xml.decode("utf-8")
    if "<drawing" in text or "<legacyDrawing" in text:
        return sheet_xml

    # openpyxl-generated sheets always have relationships (tables), but be defensive.
    if "xmlns:r=" not in text:
        text = text.replace(
            "<worksheet ",
            f'<worksheet xmlns:r="{R_NS}" ',
            1,
        )

    # Some generators may omit xmlns:r even though we inject r:id. Declare it locally to avoid Excel repair:
    # "undeclared prefix" at /xl/worksheets/sheet*.xml.
    inject = (
        f'<drawing xmlns:r="{R_NS}" r:id="{drawing_rid}"/>'
        f'<legacyDrawing xmlns:r="{R_NS}" r:id="{vml_rid}"/>'
        f'<legacyDrawingHF xmlns:r="{R_NS}" r:id="{vmlhf_rid}"/>'
    )

    idx = text.rfind("<tableParts")
    if idx != -1:
        text = text[:idx] + inject + text[idx:]
        return text.encode("utf-8")

    idx = text.rfind("</worksheet>")
    if idx == -1:
        return sheet_xml
    text = text[:idx] + inject + text[idx:]
    return text.encode("utf-8")


def _ensure_sheet_rels_has_drawings(
    *,
    rels_xml: bytes,
    drawing_target: str,
    vml_target: str,
    vmlhf_target: str,
) -> tuple[bytes, dict[str, str]]:
    """
    Ensures a sheet rels file contains relationships to:
      - drawing1.xml
      - vmlDrawing1.vml
      - vmlDrawing2.vml
    Returns (updated_xml, rId_map).
    """
    ET.register_namespace("", REL_NS)
    root = ET.fromstring(rels_xml)

    existing_ids = set()
    existing_pairs = set()
    for rel in root.findall(f"{{{REL_NS}}}Relationship"):
        existing_ids.add(rel.get("Id"))
        existing_pairs.add((rel.get("Type"), rel.get("Target")))

    def _ensure(type_: str, target: str) -> str:
        pair = (type_, target)
        if pair in existing_pairs:
            # Reuse the existing Id for this target/type.
            for rel in root.findall(f"{{{REL_NS}}}Relationship"):
                if rel.get("Type") == type_ and rel.get("Target") == target:
                    return rel.get("Id")

        rid = _pick_free_rid(existing_ids, start=2)
        ET.SubElement(
            root,
            f"{{{REL_NS}}}Relationship",
            Id=rid,
            Type=type_,
            Target=target,
        )
        existing_ids.add(rid)
        existing_pairs.add(pair)
        return rid

    rid_drawing = _ensure(
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
        drawing_target,
    )
    rid_vml = _ensure(
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing",
        vml_target,
    )
    rid_vmlhf = _ensure(
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing",
        vmlhf_target,
    )

    return (
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
        {"drawing": rid_drawing, "vml": rid_vml, "vmlhf": rid_vmlhf},
    )


def _restore_template_drawings(*, xlsx_bytes: bytes, template_path: Path) -> bytes:
    """
    openpyxl cannot preserve existing drawings/images/VML from a template workbook.
    We post-process the generated xlsx (zip) to merge the template drawing parts back in.
    """
    template_bytes = template_path.read_bytes()

    out_buf = BytesIO()
    with zipfile.ZipFile(BytesIO(xlsx_bytes), "r") as out_zip, zipfile.ZipFile(
        BytesIO(template_bytes), "r"
    ) as tmpl_zip, zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as new_zip:
        out_names = set(out_zip.namelist())

        # Template drawing parts (single-sheet) that we'll reuse per output sheet.
        tmpl_drawing_xml = tmpl_zip.read("xl/drawings/drawing1.xml")
        tmpl_drawing_rels = tmpl_zip.read("xl/drawings/_rels/drawing1.xml.rels")
        tmpl_vml1 = tmpl_zip.read("xl/drawings/vmlDrawing1.vml")
        tmpl_vml1_rels = tmpl_zip.read("xl/drawings/_rels/vmlDrawing1.vml.rels")
        tmpl_vml2 = tmpl_zip.read("xl/drawings/vmlDrawing2.vml")
        tmpl_vml2_rels = tmpl_zip.read("xl/drawings/_rels/vmlDrawing2.vml.rels")

        # Determine the content-type used for drawing xml so we can add overrides for drawing2+.
        drawing_ct = None
        try:
            ct_root = ET.fromstring(tmpl_zip.read("[Content_Types].xml"))
            ct_ns = ct_root.tag.split("}")[0].strip("{") if "}" in ct_root.tag else ""
            for o in ct_root.findall(f"{{{ct_ns}}}Override"):
                if o.get("PartName") == "/xl/drawings/drawing1.xml":
                    drawing_ct = o.get("ContentType")
                    break
        except Exception:
            drawing_ct = None

        extra_overrides: dict[str, str] = {}

        # Precompute relationship IDs per sheet so we can inject matching r:ids into sheet XML.
        sheet_rid_map: dict[str, dict[str, str]] = {}
        for name in out_names:
            if not (name.startswith("xl/worksheets/_rels/sheet") and name.endswith(".xml.rels")):
                continue
            sheet_xml_name = name.replace("xl/worksheets/_rels/", "xl/worksheets/").replace(".rels", "")
            try:
                sheet_num = int(Path(sheet_xml_name).stem.replace("sheet", ""))
            except Exception:
                sheet_num = 1

            updated_rels, rid_map = _ensure_sheet_rels_has_drawings(
                rels_xml=out_zip.read(name),
                drawing_target=f"../drawings/drawing{sheet_num}.xml",
                vml_target=f"../drawings/vmlDrawing{(sheet_num * 2) - 1}.vml",
                vmlhf_target=f"../drawings/vmlDrawing{sheet_num * 2}.vml",
            )
            sheet_rid_map[sheet_xml_name] = rid_map
            new_zip.writestr(name, updated_rels)

            # Each output sheet gets its own drawing/vml parts to avoid Excel repair prompts.
            drawing_name = f"xl/drawings/drawing{sheet_num}.xml"
            drawing_rels_name = f"xl/drawings/_rels/drawing{sheet_num}.xml.rels"
            vml_name = f"xl/drawings/vmlDrawing{(sheet_num * 2) - 1}.vml"
            vml_rels_name = f"xl/drawings/_rels/vmlDrawing{(sheet_num * 2) - 1}.vml.rels"
            vmlhf_name = f"xl/drawings/vmlDrawing{sheet_num * 2}.vml"
            vmlhf_rels_name = f"xl/drawings/_rels/vmlDrawing{sheet_num * 2}.vml.rels"

            if drawing_name not in out_names:
                new_zip.writestr(drawing_name, tmpl_drawing_xml)
            if drawing_rels_name not in out_names:
                new_zip.writestr(drawing_rels_name, tmpl_drawing_rels)
            if vml_name not in out_names:
                new_zip.writestr(vml_name, tmpl_vml1)
            if vml_rels_name not in out_names:
                new_zip.writestr(vml_rels_name, tmpl_vml1_rels)
            if vmlhf_name not in out_names:
                new_zip.writestr(vmlhf_name, tmpl_vml2)
            if vmlhf_rels_name not in out_names:
                new_zip.writestr(vmlhf_rels_name, tmpl_vml2_rels)

            if drawing_ct:
                extra_overrides[f"/xl/drawings/drawing{sheet_num}.xml"] = drawing_ct

        # Copy everything else, patching sheet XML and Content_Types as needed.
        for info in out_zip.infolist():
            name = info.filename
            if name in new_zip.namelist():
                continue

            data = out_zip.read(name)

            if name == "[Content_Types].xml":
                data = _merge_content_types(
                    out_xml=data,
                    template_xml=tmpl_zip.read(name),
                    extra_overrides=extra_overrides,
                )

            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                rid_map = sheet_rid_map.get(name)
                if rid_map:
                    data = _ensure_sheet_has_drawings(
                        sheet_xml=data,
                        drawing_rid=rid_map["drawing"],
                        vml_rid=rid_map["vml"],
                        vmlhf_rid=rid_map["vmlhf"],
                    )

            new_zip.writestr(info, data)

        # Copy the template media parts (and printer settings) into the output if missing.
        for name in tmpl_zip.namelist():
            if not (
                name.startswith("xl/media/")
                or name.startswith("xl/printerSettings/")
            ):
                continue
            if name in out_names:
                continue
            new_zip.writestr(name, tmpl_zip.read(name))

    out_buf.seek(0)
    return out_buf.getvalue()


def _normalize_label(value: str) -> str:
    # Normalize labels like "발  주  일  :" -> "발주일:" for robust matching across templates.
    return "".join(str(value or "").split())


def _find_first_cell(ws, text: str):
    for row in ws.iter_rows():
        for cell in row:
            if (cell.value or "") == text:
                return cell
    return None


def _find_cell_by_normalized(ws, normalized_text: str):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and _normalize_label(cell.value) == normalized_text:
                return cell
    return None


def _write_purchase_dates(ws, order_date: datetime, due_date: date | None = None):
    # Template labels live on row 4: "발 주 일 :" at A4 and "납  기  일 :" at D4.
    # We write values to the immediate next cells (B4 and E4).
    order_label = _find_cell_by_normalized(ws, "발주일:")
    due_label = _find_cell_by_normalized(ws, "납기일:")

    order_cell = ws.cell(row=order_label.row, column=order_label.column + 1) if order_label else ws["B3"]
    due_cell = ws.cell(row=due_label.row, column=due_label.column + 1) if due_label else ws["B4"]

    order_cell.value = order_date.date().isoformat()
    due_cell.value = due_date.isoformat() if due_date else None


def _clear_item_rows(ws, start_row: int, end_row: int):
    for r in range(start_row, end_row + 1):
        for c in range(1, 12):  # A..K
            ws.cell(r, c).value = None


def _get_bom_children(db: Session, parent_code: str):
    return db.query(models.BOM).filter(models.BOM.parent_code == parent_code).all()


def _build_row(
    *,
    part_or_finished: Optional[models.Product],
    fallback_code: str,
    finished_qty: int | None,
    part_qty: int | None,
    note: str = "",
):
    old_code = part_or_finished.old_code if part_or_finished and part_or_finished.old_code else ""
    new_code = part_or_finished.new_code if part_or_finished else fallback_code
    drawing_number = (part_or_finished.drawing_number or "").strip() if part_or_finished else ""

    return {
        "code": old_code or new_code,
        "drawing_number": drawing_number or new_code,
        "name": part_or_finished.name if part_or_finished else "",
        "spec": part_or_finished.spec if part_or_finished else "",
        "material": part_or_finished.material if part_or_finished else "",
        "finished_qty": finished_qty if finished_qty is not None else "",
        "part_qty": part_qty if part_qty is not None else "",
        "heat_treatment": part_or_finished.heat_treatment if part_or_finished else "",
        "welding": part_or_finished.welding if part_or_finished else "",
        "plating": part_or_finished.plating if part_or_finished else "",
        "note": note,
    }


def _expand_orders_for_template(db: Session, orders: list[models.PurchaseOrder]):
    rows = []

    for order in orders:
        product = db.query(models.Product).filter(
            models.Product.new_code == order.product_code
        ).first()

        qty = int(order.quantity or 0)
        product_type = (product.type or "").upper() if product else ""

        if product and product_type == "FINISHED":
            rows.append(
                _build_row(
                    part_or_finished=product,
                    fallback_code=order.product_code,
                    finished_qty=qty,
                    part_qty=None,
                )
            )

            boms = _get_bom_children(db, product.new_code)
            if not boms:
                rows[-1]["note"] = "BOM 없음"
                continue

            for bom in boms:
                child_product = db.query(models.Product).filter(
                    models.Product.new_code == bom.child_code
                ).first()
                child_qty = int(bom.quantity or 0) * qty
                rows.append(
                    _build_row(
                        part_or_finished=child_product,
                        fallback_code=bom.child_code,
                        finished_qty=None,
                        part_qty=child_qty,
                    )
                )
        else:
            # Default: treat as PART purchase order.
            rows.append(
                _build_row(
                    part_or_finished=product,
                    fallback_code=order.product_code,
                    finished_qty=None,
                    part_qty=qty,
                )
            )

    return rows


def _render_expanded_rows(ws, expanded_rows: list[dict]):
    # Template header row is 5, items start at 6.
    start_row = 6

    # Find footer row ("※ 비고 : ") to know how many lines are available.
    footer_cell = _find_first_cell(ws, "※ 비고 : ")
    footer_row = footer_cell.row if footer_cell else (start_row + 28)
    last_item_row = max(start_row, footer_row - 2)
    capacity = last_item_row - start_row + 1

    _clear_item_rows(ws, start_row, last_item_row)

    for idx, entry in enumerate(expanded_rows[:capacity]):
        row = start_row + idx
        ws.cell(row, 1).value = entry["code"]  # 품번
        ws.cell(row, 2).value = entry["drawing_number"]  # 도번
        ws.cell(row, 3).value = entry["name"]  # 품명
        ws.cell(row, 4).value = entry["spec"]  # 규격
        ws.cell(row, 5).value = entry["material"]  # 재질
        ws.cell(row, 6).value = entry["finished_qty"]  # 완제품수량
        ws.cell(row, 7).value = entry["part_qty"]  # 단품 수량
        ws.cell(row, 8).value = entry["heat_treatment"]  # 열처리
        ws.cell(row, 9).value = entry["welding"]  # 용접
        ws.cell(row, 10).value = entry["plating"]  # 도금
        ws.cell(row, 11).value = entry["note"]  # 비고

    remaining = expanded_rows[capacity:]
    return remaining


def _render_items(ws, orders: list[models.PurchaseOrder], db: Session):
    expanded_rows = _expand_orders_for_template(db, orders)
    return _render_expanded_rows(ws, expanded_rows)

def resolve_product(db: Session, input_code: str):
    product = db.query(models.Product).filter(
        models.Product.new_code == input_code
    ).first()

    if not product:
        product = db.query(models.Product).filter(
            models.Product.old_code == input_code
        ).first()

    return product

def update_batch_status(db: Session, batch_id: int):
    if not batch_id:
        return

    items = db.query(models.PurchaseOrder).filter(
        models.PurchaseOrder.batch_id == batch_id
    ).all()

    if not items:
        batch = db.query(models.PurchaseOrderBatch).filter(
            models.PurchaseOrderBatch.id == batch_id
        ).first()
        if batch:
            db.delete(batch)
            db.commit()
        return

    if all(i.status == "DONE" for i in items):
        status = "DONE"
    elif any((i.received_quantity or 0) > 0 for i in items):
        status = "PARTIAL"
    else:
        status = "WAIT"

    batch = db.query(models.PurchaseOrderBatch).filter(
        models.PurchaseOrderBatch.id == batch_id
    ).first()
    if batch:
        batch.status = status
        db.commit()


@router.get("/receipts")
def get_purchase_receipts(db: Session = Depends(get_db)):
    receipts = db.query(models.PurchaseOrderReceipt).all()
    return [
        {
            "id": r.id,
            "purchase_order_id": r.purchase_order_id,
            "quantity": r.quantity,
            "created_at": r.created_at
        }
        for r in receipts
    ]


@router.put("/receipts/{receipt_id}")
def update_receipt(receipt_id: int, data: schemas.PurchaseReceiptUpdate, db: Session = Depends(get_db)):
    receipt = db.query(models.PurchaseOrderReceipt).filter(
        models.PurchaseOrderReceipt.id == receipt_id
    ).first()
    if not receipt:
        raise Exception("입고 내역 없음")

    order = db.query(models.PurchaseOrder).filter(
        models.PurchaseOrder.id == receipt.purchase_order_id
    ).first()
    if not order:
        raise Exception("발주 없음")

    new_qty = receipt.quantity
    if data.quantity is not None:
        new_qty = int(data.quantity)
        if new_qty <= 0:
            raise Exception("수량이 올바르지 않습니다")

    delta = new_qty - (receipt.quantity or 0)
    new_received = (order.received_quantity or 0) + delta

    if new_received < 0:
        raise Exception("입고 수량이 0보다 작아질 수 없습니다")
    if new_received > (order.quantity or 0):
        raise Exception("입고 수량이 주문 수량을 초과합니다")

    # 재고 반영
    if delta != 0:
        product = db.query(models.Product).filter(
            models.Product.new_code == order.product_code
        ).first()
        if product:
            product.quantity = (product.quantity or 0) + delta

        db.add(models.Transaction(
            product_code=order.product_code,
            quantity=abs(delta),
            type="IN" if delta > 0 else "OUT",
            reason="PURCHASE_EDIT"
        ))

    receipt.quantity = new_qty

    if data.created_at:
        try:
            receipt.created_at = datetime.fromisoformat(data.created_at.replace("Z", "+00:00"))
        except Exception:
            pass

    order.received_quantity = new_received
    if new_received == 0:
        order.status = "WAIT"
    elif new_received >= (order.quantity or 0):
        order.status = "DONE"
    else:
        order.status = "PARTIAL"

    db.commit()
    update_batch_status(db, order.batch_id)

    return {"message": "수정 완료"}


@router.delete("/receipts/{receipt_id}")
def delete_receipt(receipt_id: int, db: Session = Depends(get_db)):
    receipt = db.query(models.PurchaseOrderReceipt).filter(
        models.PurchaseOrderReceipt.id == receipt_id
    ).first()
    if not receipt:
        raise Exception("입고 내역 없음")

    order = db.query(models.PurchaseOrder).filter(
        models.PurchaseOrder.id == receipt.purchase_order_id
    ).first()
    if not order:
        raise Exception("발주 없음")

    qty = receipt.quantity or 0
    new_received = (order.received_quantity or 0) - qty
    if new_received < 0:
        raise Exception("입고 수량이 0보다 작아질 수 없습니다")

    product = db.query(models.Product).filter(
        models.Product.new_code == order.product_code
    ).first()
    if product:
        product.quantity = (product.quantity or 0) - qty

    db.add(models.Transaction(
        product_code=order.product_code,
        quantity=qty,
        type="OUT",
        reason="PURCHASE_DELETE"
    ))

    order.received_quantity = new_received
    if new_received == 0:
        order.status = "WAIT"
    elif new_received >= (order.quantity or 0):
        order.status = "DONE"
    else:
        order.status = "PARTIAL"

    db.delete(receipt)
    db.commit()
    update_batch_status(db, order.batch_id)

    return {"message": "삭제 완료"}

@router.post("/batch")
def create_purchase_batch(data: schemas.PurchaseOrderBatchCreate, db: Session = Depends(get_db)):
    company = (data.company or "").strip()
    if not company:
        raise Exception("납품처가 필요합니다")

    if not data.items:
        raise Exception("발주 항목이 없습니다")

    parsed_due: date | None = None
    if data.due_date:
        try:
            parsed_due = date.fromisoformat(str(data.due_date).strip())
        except Exception:
            raise Exception("납기일 형식이 올바르지 않습니다. YYYY-MM-DD 형태로 입력해주세요.")

    batch = models.PurchaseOrderBatch(company=company, due_date=parsed_due)
    db.add(batch)
    db.commit()
    db.refresh(batch)

    created = []

    for item in data.items:
        input_code = item.product_code
        product = resolve_product(db, input_code)

        if not product:
            raise Exception("존재하지 않는 품번입니다")

        real_code = product.new_code

        db_order = models.PurchaseOrder(
            batch_id=batch.id,
            product_code=real_code,
            quantity=item.quantity,
            received_quantity=0,
            company=company
        )
        db.add(db_order)
        created.append(db_order)

    db.commit()
    return {"batch_id": batch.id, "count": len(created)}


@router.get("/batch/{batch_id}/xlsx")
def export_purchase_batch_xlsx(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(models.PurchaseOrderBatch).filter(
        models.PurchaseOrderBatch.id == batch_id
    ).first()
    if not batch:
        raise HTTPException(status_code=404, detail="발주 배치 없음")

    orders = db.query(models.PurchaseOrder).filter(
        models.PurchaseOrder.batch_id == batch_id
    ).all()
    if not orders:
        raise HTTPException(status_code=404, detail="발주 항목 없음")

    if not TEMPLATE_PATH.exists():
        raise HTTPException(status_code=500, detail="발주서 템플릿 파일이 없습니다")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    template_ws = wb.active

    # Keep a pristine copy for pagination.
    pristine = wb.copy_worksheet(template_ws)
    pristine.title = "__template__"
    pristine.sheet_state = "hidden"

    created_at = batch.created_at or datetime.utcnow()
    due_at = batch.due_date

    # First page uses the original active sheet.
    base_ws = template_ws
    base_ws.title = "발주서"
    _write_purchase_dates(base_ws, created_at, due_at)

    remaining = _render_items(base_ws, orders, db)
    page = 2
    while remaining:
        ws = wb.copy_worksheet(pristine)
        ws.title = f"발주서({page})"
        _write_purchase_dates(ws, created_at, due_at)
        remaining = _render_expanded_rows(ws, remaining)
        page += 1

    # Drop the hidden template before exporting.
    wb.remove(pristine)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # openpyxl drops template drawings (approval box / logo). Restore them from the original template zip.
    try:
        restored = _restore_template_drawings(xlsx_bytes=output.getvalue(), template_path=TEMPLATE_PATH)
        output = BytesIO(restored)
        output.seek(0)
    except Exception:
        # If restoration fails, still return a usable spreadsheet (without images).
        output.seek(0)

    # Prefer ASCII filename for broad browser compatibility; keep UTF-8 variant too.
    filename_ascii = f"purchase_order_{batch_id}.xlsx"
    filename_utf8 = f"발주서_{batch_id}.xlsx"
    headers = {
        "Content-Disposition": (
            f'attachment; filename=\"{filename_ascii}\"; '
            f"filename*=UTF-8''{quote(filename_utf8)}"
        )
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.put("/{order_id}")
def update_purchase_order(order_id: int, data: schemas.PurchaseOrderUpdate, db: Session = Depends(get_db)):
    order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == order_id).first()
    if not order:
        raise Exception("발주 없음")

    new_company = (data.company or "").strip() if data.company is not None else order.company
    new_qty = order.quantity
    if data.quantity is not None:
        new_qty = int(data.quantity)
        if new_qty <= 0:
            raise Exception("수량이 올바르지 않습니다")
        if (order.received_quantity or 0) > new_qty:
            raise Exception("이미 입고된 수량보다 작게 수정할 수 없습니다")

    new_code = order.product_code
    if data.product_code:
        product = resolve_product(db, data.product_code)
        if not product:
            raise Exception("존재하지 않는 품번입니다")
        new_code = product.new_code

    # 제품 변경 시 재고 이동 (이미 입고된 수량 반영)
    if new_code != order.product_code and (order.received_quantity or 0) > 0:
        qty = order.received_quantity or 0
        old_product = db.query(models.Product).filter(
            models.Product.new_code == order.product_code
        ).first()
        new_product = db.query(models.Product).filter(
            models.Product.new_code == new_code
        ).first()

        if old_product:
            old_product.quantity = (old_product.quantity or 0) - qty
        if new_product:
            new_product.quantity = (new_product.quantity or 0) + qty

        db.add(models.Transaction(
            product_code=order.product_code,
            quantity=qty,
            type="OUT",
            reason="PURCHASE_EDIT"
        ))
        db.add(models.Transaction(
            product_code=new_code,
            quantity=qty,
            type="IN",
            reason="PURCHASE_EDIT"
        ))

    order.product_code = new_code
    order.quantity = new_qty
    order.company = new_company

    # 상태 재계산
    if (order.received_quantity or 0) == 0:
        order.status = "WAIT"
    elif (order.received_quantity or 0) >= new_qty:
        order.status = "DONE"
    else:
        order.status = "PARTIAL"

    # 배치 회사 동기화
    if order.batch_id and data.company is not None:
        batch = db.query(models.PurchaseOrderBatch).filter(
            models.PurchaseOrderBatch.id == order.batch_id
        ).first()
        if batch:
            batch.company = new_company
        db.query(models.PurchaseOrder).filter(
            models.PurchaseOrder.batch_id == order.batch_id
        ).update({"company": new_company})

    db.commit()
    update_batch_status(db, order.batch_id)

    return {"message": "수정 완료"}


@router.post("/")
def create_purchase_order(order: schemas.PurchaseOrderCreate, db: Session = Depends(get_db)):
    return create_purchase_batch(
        schemas.PurchaseOrderBatchCreate(
            company=order.company,
            items=[schemas.PurchaseOrderItemCreate(
                product_code=order.product_code,
                quantity=order.quantity
            )]
        ),
        db
    )


@router.get("/")
def get_purchase_orders(db: Session = Depends(get_db)):
    orders = db.query(models.PurchaseOrder).all()

    result = []

    for o in orders:
        product = db.query(models.Product).filter(
            models.Product.new_code == o.product_code
        ).first()
        batch = None
        if o.batch_id:
            batch = db.query(models.PurchaseOrderBatch).filter(
                models.PurchaseOrderBatch.id == o.batch_id
            ).first()

        result.append({
            "id": o.id,
            "batch_id": o.batch_id,
            "product_code": o.product_code,
            "product_name": product.name if product else "",
            "quantity": o.quantity,
            "received_quantity": o.received_quantity or 0,
            "company": o.company,
            "status": o.status,
            "created_at": o.created_at,
            "batch_created_at": batch.created_at if batch else o.created_at,
            "batch_company": batch.company if batch else o.company,
            "batch_status": batch.status if batch else o.status
        })

    return result


@router.post("/receive/{order_id}")
def receive_purchase(order_id: int, data: schemas.PurchaseReceive, db: Session = Depends(get_db)):
    order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == order_id).first()

    if not order:
        raise Exception("주문 없음")

    if order.status == "DONE":
        raise Exception("이미 완료")

    qty = int(data.quantity or 0)
    if qty <= 0:
        raise Exception("입고 수량이 올바르지 않습니다")

    remaining = (order.quantity or 0) - (order.received_quantity or 0)
    if qty > remaining:
        raise Exception("입고 수량이 잔여 수량보다 많습니다")

    # 재고 반영
    product = db.query(models.Product).filter(
        models.Product.new_code == order.product_code
    ).first()
    if product:
        product.quantity = (product.quantity or 0) + qty

    order.received_quantity = (order.received_quantity or 0) + qty

    if order.received_quantity >= order.quantity:
        order.status = "DONE"
    else:
        order.status = "PARTIAL"

    db.add(models.PurchaseOrderReceipt(
        purchase_order_id=order.id,
        quantity=qty
    ))

    db.add(models.Transaction(
        product_code=order.product_code,
        quantity=qty,
        type="IN",
        reason="PURCHASE_IN"
    ))

    db.commit()
    update_batch_status(db, order.batch_id)

    return {"message": "입고 완료", "received_quantity": order.received_quantity, "status": order.status}


@router.post("/receive-all/{order_id}")
def receive_all(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == order_id).first()

    if not order:
        raise Exception("주문 없음")

    if order.status == "DONE":
        raise Exception("이미 완료")

    remaining = (order.quantity or 0) - (order.received_quantity or 0)
    if remaining <= 0:
        raise Exception("잔여 수량 없음")

    # 재고 반영
    product = db.query(models.Product).filter(
        models.Product.new_code == order.product_code
    ).first()
    if product:
        product.quantity = (product.quantity or 0) + remaining

    order.received_quantity = (order.received_quantity or 0) + remaining
    order.status = "DONE"

    db.add(models.PurchaseOrderReceipt(
        purchase_order_id=order.id,
        quantity=remaining
    ))

    db.add(models.Transaction(
        product_code=order.product_code,
        quantity=remaining,
        type="IN",
        reason="PURCHASE_IN"
    ))

    db.commit()
    update_batch_status(db, order.batch_id)

    return {"message": "전체 입고 완료", "received_quantity": order.received_quantity, "status": order.status}


@router.post("/complete/{order_id}")
def complete_purchase(order_id: int, db: Session = Depends(get_db)):
    return receive_all(order_id, db)


@router.post("/undo/{order_id}")
def undo_purchase(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == order_id).first()

    if not order:
        raise Exception("주문 없음")

    if order.status != "DONE":
        raise Exception("완료 상태만 취소 가능")

    order.status = "WAIT"
    db.commit()
    update_batch_status(db, order.batch_id)

    return {"message": "취소 완료"}


@router.delete("/{order_id}")
def delete_purchase(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == order_id).first()

    if not order:
        raise Exception("주문 없음")

    # 기존 입고 내역이 있으면 재고 되돌림 + 입고 내역 삭제
    receipts = db.query(models.PurchaseOrderReceipt).filter(
        models.PurchaseOrderReceipt.purchase_order_id == order.id
    ).all()
    if receipts:
        total = sum((r.quantity or 0) for r in receipts)
        product = db.query(models.Product).filter(
            models.Product.new_code == order.product_code
        ).first()
        if product:
            product.quantity = (product.quantity or 0) - total

        db.add(models.Transaction(
            product_code=order.product_code,
            quantity=total,
            type="OUT",
            reason="PURCHASE_DELETE"
        ))

        for r in receipts:
            db.delete(r)

    batch_id = order.batch_id
    db.delete(order)
    db.commit()
    update_batch_status(db, batch_id)

    return {"message": "삭제 완료"}


@router.delete("/batch/{batch_id}")
def delete_purchase_batch(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(models.PurchaseOrderBatch).filter(
        models.PurchaseOrderBatch.id == batch_id
    ).first()
    if not batch:
        raise HTTPException(status_code=404, detail="발주 배치 없음")

    orders = db.query(models.PurchaseOrder).filter(
        models.PurchaseOrder.batch_id == batch_id
    ).all()

    deleted_orders = 0
    deleted_receipts = 0

    for order in orders:
        receipts = db.query(models.PurchaseOrderReceipt).filter(
            models.PurchaseOrderReceipt.purchase_order_id == order.id
        ).all()

        if receipts:
            total = sum((r.quantity or 0) for r in receipts)
            if total:
                product = db.query(models.Product).filter(
                    models.Product.new_code == order.product_code
                ).first()
                if product:
                    product.quantity = (product.quantity or 0) - total

                db.add(models.Transaction(
                    product_code=order.product_code,
                    quantity=total,
                    type="OUT",
                    reason="PURCHASE_BATCH_DELETE"
                ))

            for r in receipts:
                db.delete(r)
                deleted_receipts += 1

        db.delete(order)
        deleted_orders += 1

    db.delete(batch)
    db.commit()

    return {
        "message": "삭제 완료",
        "batch_id": batch_id,
        "deleted_orders": deleted_orders,
        "deleted_receipts": deleted_receipts,
    }

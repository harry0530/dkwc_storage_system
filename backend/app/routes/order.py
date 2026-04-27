from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, schemas
from datetime import date, datetime, timezone
from zoneinfo import ZoneInfo
from urllib.parse import quote
from io import BytesIO
from typing import Optional
import openpyxl
from openpyxl.chart import BarChart, Reference
import os
import smtplib
from email.mime.text import MIMEText

router = APIRouter(prefix="/orders", tags=["Orders"])


def _is_finished_product(db: Session, product) -> bool:
    # Consider FINISHED if explicitly typed, or if BOM rows exist (legacy support)
    if not product:
        return False
    if (getattr(product, "type", "") or "").strip().upper() == "FINISHED":
        return True
    bom = db.query(models.BOM).filter(models.BOM.parent_code == product.new_code).first()
    return bom is not None


# ⭐ 주문 생성 (alias + 우리품번 + 검증 포함)
@router.post("/")
def create_order(order: schemas.OrderCreate, db: Session = Depends(get_db)):
    company = order.company
    input_code = order.product_code

    # 신품번 우선
    product = db.query(models.Product).filter(
        models.Product.new_code == input_code
    ).first()

    # 구품번도 허용
    if not product:
        product = db.query(models.Product).filter(
            models.Product.old_code == input_code
        ).first()

    if not product:
        raise Exception("존재하지 않는 품번입니다")

    if not _is_finished_product(db, product):
        raise Exception("존재하지 않는 완제품 품번입니다")

    real_code = product.new_code

    db_order = models.Order(
        product_code=real_code,
        quantity=order.quantity,
        company=company
    )

    db.add(db_order)
    db.commit()
    db.refresh(db_order)

    return db_order


# ⭐ 주문 조회 (제품명 포함)
@router.get("/")
def get_orders(db: Session = Depends(get_db)):
    orders = db.query(models.Order).all()

    result = []

    for o in orders:
        product = db.query(models.Product).filter(
            models.Product.new_code == o.product_code
        ).first()

        result.append({
            "id": o.id,
            "product_code": o.product_code,
            "old_code": product.old_code if product else "",
            "new_code": product.new_code if product else o.product_code,
            "product_name": product.name if product else "",
            "quantity": o.quantity,
            "company": o.company,
            "status": o.status,
            "created_at": o.created_at
        })

    return result


def _parse_ymd(value: Optional[str]) -> Optional[date]:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except Exception:
        raise HTTPException(status_code=400, detail="날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형태로 입력해 주세요.")


@router.get("/summary-xlsx")
def export_sales_summary_xlsx(
    from_: Optional[str] = Query(default=None, alias="from"),
    to: Optional[str] = Query(default=None, alias="to"),
    db: Session = Depends(get_db),
):
    """
    Export finished-product sales (orders) quantities aggregated by product, filtered by created_at date (KST).
    Returns an .xlsx with a bar chart embedded.
    """
    from_date = _parse_ymd(from_)
    to_date = _parse_ymd(to)
    if from_date and to_date and from_date > to_date:
        raise HTTPException(status_code=400, detail="시작일이 종료일보다 클 수 없습니다.")

    kst = ZoneInfo("Asia/Seoul")

    orders = db.query(models.Order).all()

    totals: dict[str, dict] = {}
    for o in orders:
        created = o.created_at
        if not created:
            continue
        if isinstance(created, datetime):
            dt = created
        else:
            continue

        # Treat naive timestamps as UTC for consistent cross-env behavior.
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        ymd = dt.astimezone(kst).date()

        if from_date and ymd < from_date:
            continue
        if to_date and ymd > to_date:
            continue

        code = (o.product_code or "").strip()
        if not code:
            continue
        key = code.lower()
        current = totals.get(key)
        if not current:
            totals[key] = {"code": code, "qty": int(o.quantity or 0)}
        else:
            current["qty"] = int(current.get("qty") or 0) + int(o.quantity or 0)

    rows = []
    for key, payload in totals.items():
        code = (payload.get("code") or "").strip()
        qty = int(payload.get("qty") or 0)
        product = db.query(models.Product).filter(models.Product.new_code == code).first()
        new_code = product.new_code if product else (code or "")
        old_code = product.old_code if product else ""
        name = product.name if product else ""
        rows.append((old_code or "", new_code or "", name or "", int(qty or 0)))

    rows.sort(key=lambda r: str(r[1] or ""))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수주집계"

    ws.append(["구품번", "신품번", "제품명", "수량"])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="F1F5F9")

    for r in rows:
        ws.append(list(r))

    # Basic column sizing
    widths = [14, 14, 36, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    ws.freeze_panes = "A2"

    # Add chart if there is data
    last_row = ws.max_row
    if last_row >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "제품별 수주 수량"
        chart.y_axis.title = "수량"
        chart.x_axis.title = "신품번"

        data = Reference(ws, min_col=4, min_row=1, max_row=last_row)  # qty
        cats = Reference(ws, min_col=2, min_row=2, max_row=last_row)  # new_code
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 26
        ws.add_chart(chart, "F2")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    tag_from = from_date.isoformat() if from_date else "all"
    tag_to = to_date.isoformat() if to_date else "all"
    filename_ascii = f"sales_summary_{tag_from}_{tag_to}.xlsx"
    filename_utf8 = f"수주집계_{tag_from}_{tag_to}.xlsx"
    headers = {
        "Content-Disposition": (
            f'attachment; filename=\"{filename_ascii}\"; '
            f"filename*=UTF-8''{quote(filename_utf8)}"
        )
    }
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ⭐ 생산
@router.post("/produce/{order_id}")
def produce(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()

    if order.status == "DONE":
        raise Exception("이미 완료")

    boms = db.query(models.BOM).filter(
        models.BOM.parent_code == order.product_code
    ).all()

    # 재고 체크
    for bom in boms:
        part = db.query(models.Product).filter(
            models.Product.new_code == bom.child_code,
            models.Product.type == "PART"
        ).first()

        required = bom.quantity * order.quantity

        if not part or part.quantity < required:
            raise Exception("재고 부족")

    # 재고 차감
    for bom in boms:
        part = db.query(models.Product).filter(
            models.Product.new_code == bom.child_code,
            models.Product.type == "PART"
        ).first()

        required = bom.quantity * order.quantity
        part.quantity -= required

        db.add(models.Transaction(
            product_code=bom.child_code,
            quantity=required,
            type="OUT",
            reason="PRODUCTION"
        ))

    order.status = "DONE"
    db.commit()

    return {"message": "생산 완료"}


# ⭐ 생산 취소 (되돌리기)
@router.post("/undo/{order_id}")
def undo(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()

    if order.status != "DONE":
        raise Exception("완료 상태만 취소 가능")

    boms = db.query(models.BOM).filter(
        models.BOM.parent_code == order.product_code
    ).all()

    # 재고 복구
    for bom in boms:
        part = db.query(models.Product).filter(
            models.Product.new_code == bom.child_code,
            models.Product.type == "PART"
        ).first()

        restore = bom.quantity * order.quantity
        part.quantity += restore

        db.add(models.Transaction(
            product_code=bom.child_code,
            quantity=restore,
            type="IN",
            reason="UNDO_PRODUCTION"
        ))

    order.status = "WAIT"
    db.commit()

    return {"message": "생산 취소 완료"}


# ⭐ 주문 삭제 (거부)
@router.delete("/{order_id}")
def delete_order(order_id: int, db: Session = Depends(get_db)):
    order = db.query(models.Order).filter(models.Order.id == order_id).first()

    if not order:
        raise Exception("주문 없음")

    if order.status != "WAIT":
        raise Exception("대기 상태만 삭제 가능")

    db.delete(order)
    db.commit()

    return {"message": "삭제 완료"}


def send_email(to_email: str, subject: str, body: str):
    host = os.getenv("SMTP_HOST", "")
    port = int(os.getenv("SMTP_PORT", "587"))
    user = os.getenv("SMTP_USER", "")
    password = os.getenv("SMTP_PASSWORD", "")
    from_email = os.getenv("SMTP_FROM", user)
    use_tls = os.getenv("SMTP_USE_TLS", "1") == "1"

    if not host or not from_email:
        raise RuntimeError("SMTP 설정이 없습니다.")

    msg = MIMEText(body, _charset="utf-8")
    msg["Subject"] = subject
    msg["From"] = from_email
    msg["To"] = to_email

    if use_tls:
        with smtplib.SMTP(host, port) as server:
            server.starttls()
            if user and password:
                server.login(user, password)
            server.send_message(msg)
    else:
        with smtplib.SMTP(host, port) as server:
            if user and password:
                server.login(user, password)
            server.send_message(msg)


@router.post("/quote-email")
def send_quote_email(data: dict, db: Session = Depends(get_db)):
    to_email = (data.get("to") or "").strip()
    subject = (data.get("subject") or "").strip()
    body = (data.get("body") or "").strip()

    if not to_email:
        raise HTTPException(status_code=400, detail="수신 이메일이 필요합니다.")
    if not subject:
        raise HTTPException(status_code=400, detail="제목이 필요합니다.")
    if not body:
        raise HTTPException(status_code=400, detail="본문이 필요합니다.")

    try:
        send_email(to_email, subject, body)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))

    return {"message": "메일 전송 완료"}
